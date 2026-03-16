import os
import io
import re
import csv
import json
import logging
import requests
from requests.exceptions import RequestException, Timeout
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
# Optional: OCR libs (not available on Render free tier — no Tesseract binary)
try:
    from PIL import Image, ImageEnhance, ImageFilter
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False


from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from datetime import datetime, timezone

# Load environment variables from the .env file
load_dotenv()

# Configure logging (replaces print statements)
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ── CRITICAL: Validate SECRET_KEY — app will refuse to start without it ──────
GROQ_API_KEY = os.getenv("GROQ_API_KEY")

secret_key = os.getenv("SECRET_KEY")
if not secret_key:
    raise RuntimeError("SECRET_KEY is not set in .env! App cannot start safely.")
app.config['SECRET_KEY'] = secret_key

# ── File upload size limit: max 5 MB ─────────────────────────────────────────
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024

# ── Database: PostgreSQL if set, else SQLite fallback ────────────────────────
db_url = os.getenv("DATABASE_URL")
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ── Rate Limiter (brute-force protection) ────────────────────────────────────
limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri="memory://")

# ── Flask-Mail (Brevo SMTP for password reset emails) ──────────────────────
app.config['MAIL_SERVER'] = 'smtp-relay.brevo.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('BREVO_USER')   # Brevo SMTP login
app.config['MAIL_PASSWORD'] = os.getenv('BREVO_KEY')    # Brevo SMTP key
app.config['MAIL_DEFAULT_SENDER'] = ('FinFlap', os.getenv('MAIL_FROM', os.getenv('BREVO_USER')))
mail = Mail(app)

# Serializer for timed password-reset tokens (30-minute expiry)
def get_reset_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

# --- DATABASE MODELS ---

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)  # 256 chars for future-proofing
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    
    expenses = db.relationship('Expense', backref='user', lazy=True, cascade='all, delete-orphan')
    goals = db.relationship('Goal', backref='user', lazy=True, cascade='all, delete-orphan')
    assets = db.relationship('Asset', backref='user', lazy=True, cascade='all, delete-orphan')
    settings = db.relationship('Settings', backref='user', uselist=False)
    custom_categories = db.relationship('CustomCategory', backref='user', lazy=True, cascade='all, delete-orphan')
    category_limits = db.relationship('CategoryLimit', backref='user', lazy=True, cascade='all, delete-orphan')
    recurring_expenses = db.relationship('RecurringExpense', backref='user', lazy=True, cascade='all, delete-orphan')

class Expense(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    date = db.Column(db.String(20), nullable=False)
    category = db.Column(db.String(50), nullable=False)

class Goal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    target_amount = db.Column(db.Float, nullable=False)
    saved_amount = db.Column(db.Float, default=0.0)
    target_date = db.Column(db.String(20), nullable=False)
    category = db.Column(db.String(50), nullable=False)

class Asset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    asset_type = db.Column(db.String(50), nullable=False)
    invested_amount = db.Column(db.Float, nullable=False)
    current_value = db.Column(db.Float, nullable=False)

class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    monthly_income = db.Column(db.Float, default=0.0)
    monthly_budget = db.Column(db.Float, default=0.0)

class CustomCategory(db.Model):
    """User-defined expense categories with emoji and color."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(50), nullable=False)
    emoji = db.Column(db.String(10), default='📌')
    color = db.Column(db.String(20), default='#6366f1')  # Indigo default

class CategoryLimit(db.Model):
    """Per-category monthly spending limit for overspending alerts."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    category = db.Column(db.String(50), nullable=False)
    monthly_limit = db.Column(db.Float, nullable=False)

class RecurringExpense(db.Model):
    """Fixed monthly costs (rent, subscriptions) that auto-log each month."""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    category = db.Column(db.String(50), nullable=False)
    day_of_month = db.Column(db.Integer, default=1)   # Day on which it recurs
    last_logged = db.Column(db.String(7), default='')  # 'YYYY-MM' of last auto-log

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# --- AUTH ROUTES ---

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('register'))
        if User.query.filter_by(email=email).first():
            flash('Email already registered.', 'danger')
            return redirect(url_for('register'))
            
        hashed_pw = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(username=username, email=email, password_hash=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        
        # Initialize empty settings for new user
        new_settings = Settings(user_id=new_user.id)
        db.session.add(new_settings)
        db.session.commit()
        
        flash('Account created successfully! Please log in.', 'success')
        return redirect(url_for('login'))
        
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")  # Brute-force protection
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and bcrypt.check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Login Unsuccessful. Please check email and password.', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

# ── Forgot Password ───────────────────────────────────────────────────────────
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        user = User.query.filter_by(email=email).first()

        # Always show success message (don't reveal if email exists)
        if user:
            try:
                s = get_reset_serializer()
                token = s.dumps(user.email, salt='password-reset')
                reset_url = url_for('reset_password', token=token, _external=True)

                msg = Message(
                    subject='🔐 FinFlap — Reset Your Password',
                    recipients=[user.email]
                )
                msg.html = f"""
                <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto;background:#0f172a;color:#e2e8f0;border-radius:16px;padding:32px;">
                    <h2 style="color:#3b82f6;margin-top:0;">🐦 FinFlap Password Reset</h2>
                    <p>Hi <strong>{user.username}</strong>,</p>
                    <p>Someone requested a password reset for your FinFlap account.
                       If this wasn't you, just ignore this email — your password won't change.</p>
                    <p>Click the button below to reset your password. This link expires in <strong>30 minutes</strong>.</p>
                    <a href="{reset_url}" style="display:inline-block;margin:16px 0;padding:14px 28px;background:#3b82f6;color:white;border-radius:10px;text-decoration:none;font-weight:bold;">
                        Reset My Password
                    </a>
                    <p style="font-size:12px;color:#64748b;">Or copy this link:<br>{reset_url}</p>
                    <hr style="border:none;border-top:1px solid #1e293b;margin:24px 0;">
                    <p style="font-size:11px;color:#64748b;">FinFlap — Smart Finance Tracker</p>
                </div>
                """
                mail.send(msg)
                logger.info("Password reset email sent to %s", email)
            except Exception as e:
                logger.error("Failed to send reset email: %s", e)
                flash('Could not send email. Please check mail configuration.', 'danger')
                return render_template('forgot_password.html')

        flash('If that email is registered, a reset link has been sent. Check your inbox (and spam)! 📬', 'success')
        return redirect(url_for('forgot_password'))

    return render_template('forgot_password.html')

# ── Reset Password (from email link) ─────────────────────────────────────────
@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    s = get_reset_serializer()
    try:
        email = s.loads(token, salt='password-reset', max_age=1800)  # 30 minutes
    except SignatureExpired:
        flash('That reset link has expired. Please request a new one.', 'danger')
        return redirect(url_for('forgot_password'))
    except BadSignature:
        flash('That reset link is invalid. Please request a new one.', 'danger')
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()
    if not user:
        flash('User not found.', 'danger')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        new_password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('reset_password.html', token=token)

        if new_password != confirm_password:
            flash('Passwords do not match.', 'danger')
            return render_template('reset_password.html', token=token)

        user.password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
        db.session.commit()
        logger.info("Password reset successful for user %s", user.email)
        flash('Password reset successfully! You can now log in with your new password. 🎉', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', token=token)


# ── PWA static file routes (served from root for correct SW scope) ──
@app.route('/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')

@app.route('/service-worker.js')
def service_worker():
    response = app.send_static_file('service-worker.js')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Cache-Control'] = 'no-cache'
    return response

@app.route('/')
@login_required
def index():
    """Serves the main HTML interface."""
    # Flask looks for this inside the 'templates' folder
    return render_template('index.html')

# --- JSON API ENDPOINTS FOR FRONTEND ---

@app.route('/api/data', methods=['GET'])
@login_required
def get_data():
    user = current_user
    expenses_data = [{"id": e.id, "description": e.description, "amount": e.amount,
                      "date": e.date, "category": e.category} for e in user.expenses]
    goals_data = [{"id": g.id, "name": g.name, "targetAmount": g.target_amount,
                   "savedAmount": g.saved_amount, "targetDate": g.target_date,
                   "category": g.category} for g in user.goals]
    assets_data = [{"id": a.id, "name": a.name, "type": a.asset_type,
                    "invested": a.invested_amount, "current": a.current_value} for a in user.assets]
    settings_data = {"income": user.settings.monthly_income, "budget": user.settings.monthly_budget} \
        if user.settings else {"income": 0, "budget": 0}
    custom_cats = [{"id": c.id, "name": c.name, "emoji": c.emoji, "color": c.color}
                   for c in user.custom_categories]
    limits_data = [{"id": l.id, "category": l.category, "limit": l.monthly_limit}
                   for l in user.category_limits]
    recurring_data = [{"id": r.id, "description": r.description, "amount": r.amount,
                       "category": r.category, "day_of_month": r.day_of_month,
                       "last_logged": r.last_logged} for r in user.recurring_expenses]

    return jsonify({
        "expenses": expenses_data,
        "goals": goals_data,
        "portfolio": assets_data,
        "settings": settings_data,
        "user": {"username": user.username, "email": user.email},
        "custom_categories": custom_cats,
        "limits": limits_data,
        "recurring": recurring_data
    })


@app.route('/api/notifications', methods=['GET'])
@login_required
def get_notifications():
    """Generates dynamic, witty financial alerts from the user's real data."""
    user = current_user
    alerts = []
    today = datetime.now()
    curr_month = today.strftime('%Y-%m')
    curr_day = today.strftime('%Y-%m-%d')

    # ── Settings ──────────────────────────────────────────────────────────
    income = user.settings.monthly_income if user.settings else 0
    budget = user.settings.monthly_budget if user.settings else 0

    # ── This month's expenses ─────────────────────────────────────────────
    this_month_exps = [e for e in user.expenses if e.date.startswith(curr_month)]
    total_spent = sum(e.amount for e in this_month_exps)

    # Category breakdown
    cat_sums = {}
    for e in this_month_exps:
        cat_sums[e.category] = cat_sums.get(e.category, 0) + e.amount

    # ── BUDGET ALERTS ─────────────────────────────────────────────────────
    if budget > 0:
        pct = (total_spent / budget) * 100
        if pct >= 100:
            alerts.append({
                "icon": "flame", "type": "danger",
                "title": "Budget Ka Katl Ho Gaya 💸",
                "msg": f"Bhai 😬 You blew your ₹{int(budget):,} budget! You spent ₹{int(total_spent):,}. Your wallet called — it's crying in a corner."
            })
        elif pct >= 80:
            alerts.append({
                "icon": "alert-triangle", "type": "warning",
                "title": "Budget Warning ⚠️",
                "msg": f"{pct:.0f}% of your budget is gone ({int(total_spent):,}/₹{int(budget):,}). At this rate... daal-roti days are near."
            })
        elif pct >= 50:
            alerts.append({
                "icon": "pie-chart", "type": "info",
                "title": "Halfway Through Budget 📊",
                "msg": f"You're at {pct:.0f}% of your monthly budget. Not bad, not great. The month's watching you."
            })

    # ── BIG SPEND ALERT ───────────────────────────────────────────────────
    if budget > 0 and this_month_exps:
        biggest = max(this_month_exps, key=lambda e: e.amount)
        if biggest.amount >= 0.2 * budget:
            alerts.append({
                "icon": "zap", "type": "warning",
                "title": f"Big Spend Detected 🤑",
                "msg": f"A whopping ₹{int(biggest.amount):,} on '{biggest.description}' ({biggest.category})? Bold move. We respect it… barely."
            })

    # ── TOP CATEGORY ─────────────────────────────────────────────────────
    if cat_sums:
        top_cat = max(cat_sums, key=cat_sums.get)
        alerts.append({
            "icon": "trophy", "type": "info",
            "title": f"Certified {top_cat} Fanatic 🏆",
            "msg": f"Your #1 spend this month is '{top_cat}' at ₹{int(cat_sums[top_cat]):,}. A true connoisseur."
        })

    # ── ZERO SPENDING TODAY ───────────────────────────────────────────────
    today_exps = [e for e in user.expenses if e.date == curr_day]
    if not today_exps and this_month_exps:
        alerts.append({
            "icon": "star", "type": "success",
            "title": "Zero Spend Day 🤩",
            "msg": "You didn't spend a single rupee today. Historic. Screenshot this. Frame it."
        })

    # ── GOAL ALERTS ───────────────────────────────────────────────────────
    for g in user.goals:
        if g.target_amount <= 0:
            continue
        pct = (g.saved_amount / g.target_amount) * 100
        if pct >= 100:
            alerts.append({
                "icon": "check-circle", "type": "success",
                "title": f"Goal Achieved! 🎉",
                "msg": f"GOAT behavior! You've fully funded '{g.name}'. Time to celebrate (responsibly, of course)."
            })
        elif pct >= 90:
            alerts.append({
                "icon": "target", "type": "success",
                "title": f"Almost There! 🎯",
                "msg": f"Your '{g.name}' goal is {pct:.0f}% funded. One final push. You've got this, champion!"
            })
        elif pct >= 50:
            alerts.append({
                "icon": "trending-up", "type": "info",
                "title": f"Goal Halfway Done 💪",
                "msg": f"'{g.name}' is {pct:.0f}% funded. Keep going — your future self is rooting for you!"
            })

    # ── PORTFOLIO ALERTS ──────────────────────────────────────────────────
    if user.assets:
        total_inv = sum(a.invested_amount for a in user.assets)
        total_cur = sum(a.current_value for a in user.assets)
        diff = total_cur - total_inv
        if diff > 0:
            alerts.append({
                "icon": "trending-up", "type": "success",
                "title": "Portfolio is Pumping 📈",
                "msg": f"You're up ₹{int(diff):,} on your investments! You're basically Warren Buffett... kinda."
            })
        elif diff < 0:
            alerts.append({
                "icon": "trending-down", "type": "danger",
                "title": "Portfolio in the Red 📉",
                "msg": f"Your portfolio is down ₹{int(abs(diff)):,}. Don't panic. Markets go up. Probably. Hopefully."
            })

    # ── RECURRING EXPENSE REMINDERS ───────────────────────────────────────
    for r in user.recurring_expenses:
        if r.last_logged != curr_month:
            alerts.append({
                "icon": "repeat", "type": "warning",
                "title": f"Recurring Bill Due 🔁",
                "msg": f"Your '{r.description}' of ₹{int(r.amount):,} is due this month and hasn't been logged yet. Don't ghost it!"
            })

    # ── ALL CLEAR ─────────────────────────────────────────────────────────
    if not alerts:
        alerts.append({
            "icon": "shield-check", "type": "success",
            "title": "All Clear! 🌟",
            "msg": "No financial drama today. Your money is safe, your goals are on track, and your portfolio... exists. 10/10."
        })

    return jsonify({"notifications": alerts, "count": len(alerts)})


@app.route('/api/settings', methods=['POST'])
@login_required
def update_settings():
    data = request.json
    if not current_user.settings:
        current_user.settings = Settings(user_id=current_user.id)
        db.session.add(current_user.settings)
    current_user.settings.monthly_income = float(data.get('income', 0))
    current_user.settings.monthly_budget = float(data.get('budget', 0))
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/expense', methods=['POST'])
@login_required
def add_expense():
    data = request.json
    # ── Input Validation ──────────────────────────────────────────────────
    description = str(data.get('description', '')).strip()[:200]
    amount = float(data.get('amount', 0))
    date = str(data.get('date', '')).strip()
    category = str(data.get('category', 'Other')).strip()[:50]

    if not description:
        return jsonify({"error": "Description is required"}), 400
    if amount <= 0 or amount > 10_000_000:
        return jsonify({"error": "Invalid amount. Must be between 1 and 1,00,00,000."}), 400
    if not date:
        return jsonify({"error": "Date is required"}), 400

    new_exp = Expense(user_id=current_user.id, description=description,
                      amount=amount, date=date, category=category)
    db.session.add(new_exp)
    db.session.commit()
    return jsonify({"status": "success", "id": new_exp.id})

@app.route('/api/expense/<int:id>', methods=['DELETE'])
@login_required
def delete_expense(id):
    exp = db.get_or_404(Expense, id)
    if exp.user_id == current_user.id:
        db.session.delete(exp)
        db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/goal', methods=['POST'])
@login_required
def add_goal():
    data = request.json
    if data.get('id'):
        goal = db.session.get(Goal, data['id'])
        if goal and goal.user_id == current_user.id:
            goal.name = data.get('name', data.get('title'))
            goal.target_amount = float(data['targetAmount'])
            goal.saved_amount = float(data['savedAmount'])
            goal.target_date = data['targetDate']
            goal.category = data['category']
    else:
        goal = Goal(user_id=current_user.id, name=data.get('name', data.get('title')), 
                    target_amount=float(data['targetAmount']), 
                    saved_amount=float(data['savedAmount']), 
                    target_date=data['targetDate'], 
                    category=data['category'])
        db.session.add(goal)
    db.session.commit()
    return jsonify({"status": "success", "id": goal.id})

@app.route('/api/goal/<int:id>', methods=['DELETE'])
@login_required
def delete_goal(id):
    goal = db.get_or_404(Goal, id)
    if goal.user_id == current_user.id:
        db.session.delete(goal)
        db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/asset', methods=['POST'])
@login_required
def add_asset():
    data = request.json
    # ── Input Validation ──────────────────────────────────────────────────
    name = str(data.get('name', '')).strip()[:100]
    asset_type = str(data.get('type', 'Other')).strip()[:50]
    invested = float(data.get('invested', 0))
    current_val = float(data.get('current', 0))
    if not name:
        return jsonify({"error": "Asset name is required"}), 400
    if invested < 0 or invested > 1_000_000_000:
        return jsonify({"error": "Invalid invested amount"}), 400

    if data.get('id'):
        asset = db.session.get(Asset, data['id'])
        if asset and asset.user_id == current_user.id:
            asset.name = name
            asset.asset_type = asset_type
            asset.invested_amount = invested
            asset.current_value = current_val
    else:
        asset = Asset(user_id=current_user.id, name=name, asset_type=asset_type,
                      invested_amount=invested, current_value=current_val)
        db.session.add(asset)
    db.session.commit()
    return jsonify({"status": "success", "id": asset.id})

@app.route('/api/asset/<int:id>', methods=['DELETE'])
@login_required
def delete_asset(id):
    asset = db.get_or_404(Asset, id)
    if asset.user_id == current_user.id:
        db.session.delete(asset)
        db.session.commit()
    return jsonify({"status": "success"})


# ─────────────────────────────────────────────
# FEATURE 1 – EXPORT  (Excel / CSV)
# ─────────────────────────────────────────────

def _style_header(ws, headers, header_fill):
    """Write bold, coloured header row to a worksheet."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18


@app.route('/api/export/expenses', methods=['GET'])
@login_required
def export_expenses():
    """Download all expenses as an .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    fill = PatternFill("solid", fgColor="2563EB")
    _style_header(ws, ["Date", "Description", "Category", "Amount (₹)"], fill)

    for e in current_user.expenses:
        ws.append([e.date, e.description, e.category, e.amount])

    # Add totals row
    total = sum(e.amount for e in current_user.expenses)
    ws.append([])
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=total).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='finflap_expenses.xlsx')


@app.route('/api/export/portfolio', methods=['GET'])
@login_required
def export_portfolio():
    """Download all portfolio assets as an .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    fill = PatternFill("solid", fgColor="10B981")
    _style_header(ws, ["Asset Name", "Type", "Invested (₹)", "Current Value (₹)", "Gain/Loss (₹)", "Return %"], fill)

    for a in current_user.assets:
        gain = a.current_value - a.invested_amount
        pct = (gain / a.invested_amount * 100) if a.invested_amount else 0
        ws.append([a.name, a.asset_type, a.invested_amount, a.current_value, round(gain, 2), round(pct, 2)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='finflap_portfolio.xlsx')


@app.route('/api/export/csv', methods=['GET'])
@login_required
def export_expenses_csv():
    """Download all expenses as a .csv (for re-import)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Description", "Category", "Amount"])
    for e in current_user.expenses:
        writer.writerow([e.date, e.description, e.category, e.amount])

    output = io.BytesIO(buf.getvalue().encode('utf-8'))
    return send_file(output, mimetype='text/csv', as_attachment=True,
                     download_name='finflap_expenses.csv')


# ─────────────────────────────────────────────
# FEATURE 1 – IMPORT  (CSV upload)
# ─────────────────────────────────────────────

@app.route('/api/import/expenses', methods=['POST'])
@login_required
def import_expenses():
    """Accept a CSV upload and bulk-insert expenses."""
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files['file']
    if not f.filename.endswith('.csv'):
        return jsonify({"error": "Only .csv files are supported"}), 400

    stream = io.StringIO(f.stream.read().decode('utf-8'))
    reader = csv.DictReader(stream)

    imported = 0
    skipped = 0
    for row in reader:
        try:
            exp = Expense(
                user_id=current_user.id,
                date=row.get('Date', '').strip(),
                description=row.get('Description', 'Imported').strip(),
                category=row.get('Category', 'Other').strip(),
                amount=float(row.get('Amount', 0))
            )
            db.session.add(exp)
            imported += 1
        except Exception:
            skipped += 1

    db.session.commit()
    return jsonify({"status": "success", "imported": imported, "skipped": skipped})


# ─────────────────────────────────────────────
# FEATURE 2 – OCR RECEIPT SCANNING
# ─────────────────────────────────────────────

@app.route('/api/scan-receipt', methods=['POST'])
@login_required
def scan_receipt():
    """Accept a receipt image, run OCR with enhanced preprocessing, ask AI to extract structured Indian data."""
    from datetime import datetime as dt

    if not OCR_AVAILABLE:
        return jsonify({"error": "Receipt scanning is not available on this server. Please add expenses manually or use Voice Input."}), 503
    if 'image' not in request.files:
        return jsonify({"error": "No image uploaded"}), 400

    img_file = request.files['image']

    # ── STEP 1: Enhanced Image Preprocessing ─────────────────────────────
    try:
        image = Image.open(img_file.stream).convert('RGB')

        # Upscale small images to ~2000px wide for better OCR resolution
        w, h = image.size
        scale = max(1, 2000 // max(w, 1))
        if scale > 1:
            image = image.resize((w * scale, h * scale), Image.LANCZOS)

        # Greyscale → contrast boost → sharpen → median denoise
        image = image.convert('L')
        image = ImageEnhance.Contrast(image).enhance(2.5)
        image = ImageEnhance.Sharpness(image).enhance(3.0)
        image = image.filter(ImageFilter.MedianFilter(size=3))

        # First try: structured single-column (good for receipts)
        raw_text = pytesseract.image_to_string(image, config='--psm 4 --oem 3', lang='eng')

        # Fallback: uniform block mode if first gives too little text
        if len(raw_text.strip()) < 20:
            raw_text = pytesseract.image_to_string(image, config='--psm 6 --oem 3', lang='eng')

    except Exception as e:
        return jsonify({"error": f"OCR failed: {str(e)}"}), 500

    if not raw_text.strip():
        return jsonify({"error": "Could not read any text from the image. Please use a clearer photo."}), 422

    # ── STEP 2: Clean Common OCR Errors Before AI ────────────────────────
    cleaned = raw_text

    # Fix: ₹ / Rs misread as '2' or 'Z' before a number  e.g. "2 1200" → "RS 1200"
    cleaned = re.sub(r'\b(?:Rs\.?|INR|2(?=\s*[\d,]+)|Z(?=\s*[\d,]+))\s*', 'RS ', cleaned)

    # Remove spaces inside numbers: "1 200" → "1200", "1 00 000" → "100000"
    cleaned = re.sub(r'(\d)\s+(\d{3})\b', r'\1\2', cleaned)
    cleaned = re.sub(r'(\d)\s+(\d{3})\b', r'\1\2', cleaned)  # Run twice for "1 00 000"

    # Expand shorthand: 1.2k/1.2K → 1200, 2.5L/2.5 lakh → 250000
    cleaned = re.sub(r'(\d+\.?\d*)\s*[kK]\b', lambda m: str(int(float(m.group(1)) * 1000)), cleaned)
    cleaned = re.sub(r'(\d+\.?\d*)\s*[Ll](?:akh)?\b',
                     lambda m: str(int(float(m.group(1)) * 100000)), cleaned)

    # Fix dash-decimal: "1200-00" → "1200.00" (common on thermal printers)
    cleaned = re.sub(r'(\d+)-(\d{2})\b', r'\1.\2', cleaned)

    today_str = datetime.utcnow().strftime('%Y-%m-%d')

    # ── STEP 3: AI Parsing with Strict Instructions ───────────────────────
    ai_prompt = f"""You are an expert Indian receipt parser. Extract structured data from the OCR text below.
Reply ONLY with valid JSON — no markdown, no explanation, no extra words.

AMOUNT RULES:
- Find the TOTAL / GRAND TOTAL / AMOUNT DUE (the final amount the customer paid).
- Indian formats to handle: RS 1200 = 1200 | RS 1,200 = 1200 | RS 1,00,000 = 100000 | RS 450.50 = 450.50
- If text shows "2 1200" or "Z 1200" the amount is 1200 (2/Z = misread ₹ symbol).
- Strip all currency symbols, commas, spaces — return a plain float.
- NEVER return 0 if any number greater than 10 is present.

DATE RULES:
- Accepted input formats: DD/MM/YYYY, MM/DD/YYYY, DD-MM-YYYY, DD.MM.YYYY,
  "11 Mar 2025", "March 11, 2025", "11 March 2025", "11-03-25".
- Always output date as YYYY-MM-DD.
- If no date found, output: {today_str}

DESCRIPTION: Use the store/merchant name from the top of the receipt. If unclear, use the main item.

CATEGORY (choose exactly one): Food & Dining | Shopping | Entertainment | Transport | Utilities | Healthcare | Education | Other

Output exactly:
{{"description": "...", "amount": 0.0, "category": "...", "date": "YYYY-MM-DD"}}

OCR TEXT:
---
{cleaned[:3000]}
---"""

    try:
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [
                {"role": "system", "content": "You parse Indian receipts. Reply ONLY with valid JSON. No markdown. No extra text."},
                {"role": "user", "content": ai_prompt}
            ],
            "temperature": 0.0,   # Deterministic — minimises hallucination
            "max_tokens": 150
        }
        resp = requests.post(url, headers=headers, json=payload, timeout=15)
        resp.raise_for_status()
        ai_text = resp.json()['choices'][0]['message']['content'].strip()

        # Strip markdown fences if model ignores instructions
        ai_text = re.sub(r'```(?:json)?\s*|\s*```', '', ai_text).strip()

        # Extract just the JSON object in case there's surrounding text
        match = re.search(r'\{.*?\}', ai_text, re.DOTALL)
        if match:
            ai_text = match.group(0)

        parsed = json.loads(ai_text)

        # ── Post-process amount ──────────────────────────────────────────
        if 'amount' in parsed:
            amt_str = str(parsed['amount'])
            amt_str = re.sub(r'[₹,\sRSrs]', '', amt_str)   # Remove ₹ Rs commas spaces
            amt_str = re.sub(r'[^\d.]', '', amt_str)         # Keep only digits and dot
            # Handle multiple dots: keep only last decimal part
            parts = amt_str.split('.')
            if len(parts) > 2:
                amt_str = ''.join(parts[:-1]) + '.' + parts[-1]
            parsed['amount'] = float(amt_str) if amt_str and amt_str != '.' else 0.0

        # ── Post-process date ────────────────────────────────────────────
        if 'date' in parsed:
            raw_date = str(parsed['date']).strip()
            date_formats = [
                '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y',
                '%d.%m.%Y', '%d %b %Y', '%B %d, %Y', '%d %B %Y',
                '%d-%m-%y', '%d/%m/%y', '%y-%m-%d'
            ]
            for fmt in date_formats:
                try:
                    parsed['date'] = dt.strptime(raw_date, fmt).strftime('%Y-%m-%d')
                    break
                except ValueError:
                    continue
            else:
                parsed['date'] = today_str  # Fallback to today

        parsed['raw_text'] = raw_text[:300]
        return jsonify(parsed)

    except json.JSONDecodeError:
        return jsonify({
            "raw_text": raw_text[:300],
            "cleaned_text": cleaned[:300],
            "error": "AI returned non-JSON. Raw text attached — please fill manually."
        })
    except Exception as e:
        return jsonify({"raw_text": raw_text[:300], "error": f"AI parse failed: {str(e)}"})


@app.route('/api/parse-voice-expense', methods=['POST'])
@login_required
def parse_voice_expense():
    """Takes transcribed voice text from the frontend, uses Groq AI to extract expense details, and saves it."""
    data = request.json
    transcript = data.get('text', '').strip()
    
    if not transcript:
        return jsonify({"error": "No voice transcript provided."}), 400
        
    today_str = datetime.now().strftime('%Y-%m-%d')
    
    ai_prompt = f"""You are a financial AI assistant. Extract expense details from the following transcript.
Reply ONLY with valid JSON — no markdown, no explanation.
Today's date is {today_str}. If the user mentions a day like "yesterday", calculate the correct date relative to today.

CATEGORY MUST BE EXACTLY ONE OF: Food & Dining | Shopping | Entertainment | Transport | Utilities | Healthcare | Education | Other
AMOUNT must be a plain float number.

Output exactly:
{{"description": "...", "amount": 0.0, "category": "...", "date": "YYYY-MM-DD"}}

VOICE TRANSCRIPT:
---
{transcript}
---"""

    try:
        import re, json
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [
                {"role": "system", "content": "You are a financial assistant. Reply ONLY with JSON."},
                {"role": "user", "content": ai_prompt}
            ],
            "temperature": 0.0,
            "max_tokens": 150
        }
        
        resp = requests.post(url, headers=headers, json=payload, timeout=10)
        resp.raise_for_status()
        ai_text = resp.json()['choices'][0]['message']['content'].strip()
        
        # Strip markdown fences
        ai_text = re.sub(r'```(?:json)?\s*|\s*```', '', ai_text).strip()
        match = re.search(r'\{.*?\}', ai_text, re.DOTALL)
        if match:
            ai_text = match.group(0)
            
        parsed = json.loads(ai_text)
        
        description = parsed.get('description', 'Voice Expense')
        amount = float(parsed.get('amount', 0.0))
        category = parsed.get('category', 'Other')
        date = parsed.get('date', today_str)
        
        if amount <= 0:
            return jsonify({"error": "No valid amount found in transcript."}), 400
            
        # Save to database immediately
        new_expense = Expense(
            user_id=current_user.id,
            description=description,
            amount=amount,
            category=category,
            date=date
        )
        db.session.add(new_expense)
        db.session.commit()
        
        return jsonify({"success": True, "expense": {
            "id": new_expense.id,
            "description": description,
            "amount": amount,
            "category": category,
            "date": date
        }})
        
    except requests.exceptions.Timeout:
        return jsonify({"error": "AI service timeout."}), 504
    except Exception as e:
        return jsonify({"error": f"AI voice parsing failed: {str(e)}"}), 500


@app.route('/api/forecast', methods=['GET'])
@login_required
def get_forecast():
    """Analyze historical data and provide AI-driven forecast for next month."""
    from collections import defaultdict
    from datetime import datetime as dt
    
    expenses = Expense.query.filter_by(user_id=current_user.id).all()
    if not expenses:
        return jsonify({"error": "No historical data found. Add some expenses first."}), 400

    # Group by month and category for the last 6 months
    history = defaultdict(lambda: defaultdict(float))
    for e in expenses:
        try:
            m = e.date[:7] # YYYY-MM
            history[m][e.category] += e.amount
        except: continue

    sorted_months = sorted(history.keys(), reverse=True)[:6]
    history_summary = ""
    for m in sorted_months:
        total = sum(history[m].values())
        cats = ", ".join([f"{c}: ₹{amt}" for c, amt in history[m].items()])
        history_summary += f"- {m}: Total ₹{total} | Categories: {cats}\n"

    budget = current_user.settings.monthly_budget if current_user.settings else 0
    income = current_user.settings.monthly_income if current_user.settings else 0

    ai_prompt = f"""You are a financial analyst. Analyze these last 6 months of expenses and forecast the next month.
USER DATA:
- Historical Spending:
{history_summary}
- Monthly Income: ₹{income}
- Monthly Budget: ₹{budget}

TASKS:
1. Forecast total spending for NEXT month.
2. Provide 3 specific 'Smart Savings Tips' based on their category spending.
3. Give a 'Confidence Score' (0-100%) for your forecast.

Reply ONLY with JSON:
{{
  "forecast_amount": 0.0,
  "confidence": 0,
  "tips": ["Tip 1", "Tip 2", "Tip 3"],
  "analysis": "Brief 1-sentence analysis"
}}"""

    try:
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [
                {"role": "system", "content": "You are a financial forecaster. Reply only with valid JSON."},
                {"role": "user", "content": ai_prompt}
            ],
            "temperature": 0.2,
            "max_tokens": 300
        }
        resp = requests.post(url, headers=headers, json=payload, timeout=15)
        resp.raise_for_status()
        ai_data = resp.json()['choices'][0]['message']['content'].strip()
        
        # Clean markdown if needed
        ai_data = re.sub(r'```(?:json)?\s*|\s*```', '', ai_data).strip()
        return jsonify(json.loads(ai_data))
    except Exception as e:
        # Fallback simple linear forecast if AI fails
        avg_spend = sum([sum(history[m].values()) for m in sorted_months]) / len(sorted_months)
        return jsonify({
            "forecast_amount": round(avg_spend, 2),
            "confidence": 50,
            "tips": ["Track recurring expenses.", "Avoid impulse shopping.", "Stick to your budget."],
            "analysis": "Calculated based on 6-month historical average (AI Offline)."
        })

# ─────────────────────────────────────────────────────────
# FEATURE D — CUSTOM CATEGORIES
# ─────────────────────────────────────────────────────────

@app.route('/api/categories', methods=['GET'])
@login_required
def get_categories():
    cats = [{"id": c.id, "name": c.name, "emoji": c.emoji, "color": c.color}
            for c in current_user.custom_categories]
    return jsonify(cats)

@app.route('/api/categories', methods=['POST'])
@login_required
def add_category():
    data = request.json
    cat = CustomCategory(user_id=current_user.id, name=data['name'],
                         emoji=data.get('emoji', '📌'), color=data.get('color', '#6366f1'))
    db.session.add(cat)
    db.session.commit()
    return jsonify({"id": cat.id, "name": cat.name, "emoji": cat.emoji, "color": cat.color})

@app.route('/api/categories/<int:id>', methods=['DELETE'])
@login_required
def delete_category(id):
    cat = CustomCategory.query.get_or_404(id)
    if cat.user_id == current_user.id:
        # Move existing expenses to "Other"
        Expense.query.filter_by(user_id=current_user.id, category=cat.name).update({"category": "Other"})
        db.session.delete(cat)
        db.session.commit()
    return jsonify({"status": "success"})


# ─────────────────────────────────────────────────────────
# FEATURE B — CATEGORY SPENDING LIMITS
# ─────────────────────────────────────────────────────────

@app.route('/api/limits', methods=['GET'])
@login_required
def get_limits():
    return jsonify([{"id": l.id, "category": l.category, "limit": l.monthly_limit}
                    for l in current_user.category_limits])

@app.route('/api/limits', methods=['POST'])
@login_required
def save_limit():
    data = request.json
    # Update existing or create new
    existing = CategoryLimit.query.filter_by(user_id=current_user.id, category=data['category']).first()
    if existing:
        existing.monthly_limit = float(data['limit'])
    else:
        db.session.add(CategoryLimit(user_id=current_user.id, category=data['category'],
                                     monthly_limit=float(data['limit'])))
    db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/limits/<int:id>', methods=['DELETE'])
@login_required
def delete_limit(id):
    lim = CategoryLimit.query.get_or_404(id)
    if lim.user_id == current_user.id:
        db.session.delete(lim)
        db.session.commit()
    return jsonify({"status": "success"})


# ─────────────────────────────────────────────────────────
# FEATURE C — RECURRING EXPENSES
# ─────────────────────────────────────────────────────────

@app.route('/api/recurring', methods=['GET'])
@login_required
def get_recurring():
    return jsonify([{"id": r.id, "description": r.description, "amount": r.amount,
                     "category": r.category, "day_of_month": r.day_of_month, "last_logged": r.last_logged}
                    for r in current_user.recurring_expenses])

@app.route('/api/recurring', methods=['POST'])
@login_required
def save_recurring():
    data = request.json
    if data.get('id'):
        r = RecurringExpense.query.get(data['id'])
        if r and r.user_id == current_user.id:
            r.description = data['description']
            r.amount = float(data['amount'])
            r.category = data['category']
            r.day_of_month = int(data.get('day_of_month', 1))
    else:
        r = RecurringExpense(user_id=current_user.id, description=data['description'],
                             amount=float(data['amount']), category=data['category'],
                             day_of_month=int(data.get('day_of_month', 1)), last_logged='')
        db.session.add(r)
    db.session.commit()
    return jsonify({"status": "success", "id": r.id})

@app.route('/api/recurring/<int:id>', methods=['DELETE'])
@login_required
def delete_recurring(id):
    r = RecurringExpense.query.get_or_404(id)
    if r.user_id == current_user.id:
        db.session.delete(r)
        db.session.commit()
    return jsonify({"status": "success"})

@app.route('/api/recurring/log-due', methods=['POST'])
@login_required
def log_due_recurring():
    """Auto-log recurring expenses that are due this month and haven't been logged yet."""
    now = datetime.utcnow()
    current_month = now.strftime('%Y-%m')
    today_day = now.day
    logged = []

    for r in current_user.recurring_expenses:
        # Log if: not logged this month AND today >= the scheduled day
        if r.last_logged != current_month and today_day >= r.day_of_month:
            log_date = f"{current_month}-{r.day_of_month:02d}"
            expense = Expense(user_id=current_user.id, description=f"🔁 {r.description}",
                              amount=r.amount, date=log_date, category=r.category)
            db.session.add(expense)
            r.last_logged = current_month
            logged.append(r.description)

    db.session.commit()
    return jsonify({"status": "success", "auto_logged": logged})


# ─────────────────────────────────────────────────────────
# FEATURE 3E — UPI SMS PARSER
# ─────────────────────────────────────────────────────────

@app.route('/api/parse-sms', methods=['POST'])
@login_required
def parse_sms():
    """Parse Indian bank/UPI SMS texts to extract expense data using regex + AI."""
    import re, json
    data = request.json
    sms_text = data.get('sms', '').strip()
    if not sms_text:
        return jsonify({"error": "No SMS text provided"}), 400

    # ── Regex pre-extraction for common Indian bank SMS patterns ────────
    # Patterns: "debited Rs.500", "debited INR 1,200.50", "₹500 debited", etc.
    amount_patterns = [
        r'(?:debited|paid|spent|txn of|amount of|of)\s*(?:Rs\.?|INR|₹)?\s*([\d,]+(?:\.\d{1,2})?)',
        r'(?:Rs\.?|INR|₹)\s*([\d,]+(?:\.\d{1,2})?)\s*(?:debited|paid|charged|deducted)',
        r'(?:Rs\.?|INR|₹)\s*([\d,]+(?:\.\d{1,2})?)',
    ]
    amount = None
    for pat in amount_patterns:
        m = re.search(pat, sms_text, re.IGNORECASE)
        if m:
            amt_str = m.group(1).replace(',', '')
            try:
                amount = float(amt_str)
                break
            except ValueError:
                continue

    # Date patterns
    date_patterns = [
        r'(\d{2}[-/]\d{2}[-/]\d{2,4})',
        r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{2,4})',
    ]
    raw_date = None
    for pat in date_patterns:
        m = re.search(pat, sms_text, re.IGNORECASE)
        if m:
            raw_date = m.group(1)
            break

    today_str = datetime.utcnow().strftime('%Y-%m-%d')

    # ── AI to extract merchant name and category ─────────────────────────
    ai_prompt = f"""Parse this Indian bank/UPI SMS message and reply ONLY with valid JSON:
{{"description": "<merchant or payee name>", "amount": {amount or 0}, "category": "<Food & Dining|Shopping|Entertainment|Transport|Utilities|Healthcare|Education|Other>", "date": "{raw_date or today_str}"}}

Rules:
- description: merchant/payee/UPI ID name (e.g. "Swiggy", "Amazon Pay", "HDFC Credit Card")
- amount: {amount if amount else "extract the debited/paid amount as a float"}
- date: convert to YYYY-MM-DD, use {today_str} if not found
- Category based on merchant name

SMS: {sms_text}"""

    try:
        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [
                {"role": "system", "content": "Parse Indian bank SMS. Reply ONLY with JSON. No extra text."},
                {"role": "user", "content": ai_prompt}
            ],
            "temperature": 0.0, "max_tokens": 120
        }
        resp = requests.post(url, headers=headers, json=payload, timeout=10)
        resp.raise_for_status()
        ai_text = resp.json()['choices'][0]['message']['content'].strip()
        ai_text = re.sub(r'```(?:json)?\s*|\s*```', '', ai_text).strip()
        match = re.search(r'\{.*?\}', ai_text, re.DOTALL)
        if match:
            ai_text = match.group(0)
        parsed = json.loads(ai_text)

        # Ensure amount from regex wins if AI gives wrong value
        if amount and (not parsed.get('amount') or parsed['amount'] == 0):
            parsed['amount'] = amount

        # Normalize date
        raw_d = str(parsed.get('date', today_str))
        from datetime import datetime as dt
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d %b %Y'):
            try:
                parsed['date'] = dt.strptime(raw_d, fmt).strftime('%Y-%m-%d')
                break
            except ValueError:
                continue
        else:
            parsed['date'] = today_str

        return jsonify(parsed)
    except Exception as e:
        # Return regex-extracted data as fallback
        return jsonify({"description": "UPI Transaction", "amount": amount or 0,
                        "category": "Other", "date": today_str,
                        "note": f"AI unavailable, regex result: {str(e)}"})


@app.route('/api/chat', methods=['POST'])
@login_required
def chat_with_groq():
    """Secure endpoint for conversational UI via Groq."""
    if not GROQ_API_KEY:
         return jsonify({"error": "Server configuration error: Groq API key is missing."}), 500

    try:
        data = request.get_json()
        messages = data.get('messages')

        if not messages:
            return jsonify({"error": "No messages provided"}), 400

        # ── Safety: cap message content length to prevent abuse ────────────────
        for m in messages:
            if len(str(m.get('content', ''))) > 2000:
                return jsonify({"error": "Message too long. Please keep it under 2000 characters."}), 400

        # System prompt to enforce behavior (emojis, tables, concise)
        system_prompt = {
            "role": "system",
            "content": (
                "You are FinPal, a premier Indian financial advisor. "
                "Analyze the user's specific financial data (income, expenses, transactions, assets) and provide BRIEF, actionable advice. "
                "Avoid generic explanations; focus on data-driven optimization. Use Markdown and Emojis. "
                "ALWAYS use Indian Rupee (₹). NEVER use dummy examples or other currencies. Be expert and direct."
            )
        }
        
        # Prepend the system prompt to the user's message history
        full_conversation = [system_prompt] + messages

        url = "https://api.groq.com/openai/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "llama-3.3-70b-versatile", # Reliable, fast Groq model
            "messages": full_conversation,
            "temperature": 0.5,
            "max_tokens": 1024
        }

        # Make request to Groq
        response = requests.post(url, headers=headers, json=payload, timeout=20)
        response.raise_for_status()
        
        groq_data = response.json()
        advice_text = groq_data['choices'][0]['message']['content']
        
        return jsonify({"text": advice_text})

    except Timeout:
        logger.warning("AI request timed out.")
        return jsonify({"error": "The AI is taking too long to respond. Please try again later."}), 504
    except RequestException as e:
        logger.error("AI request network error: %s", e)
        return jsonify({"error": "Network issue connecting to the AI service."}), 502
    except Exception as e:
        logger.error("AI chat error: %s", e)
        return jsonify({"error": "I'm having trouble connecting to the AI brain right now. Please try again later."}), 500

# ── Auto-create tables on startup (runs for both Gunicorn on Render AND local) ──
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    # ── SAFE: debug mode is only on when FLASK_DEBUG=true in .env ───────────────
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, port=5000)
